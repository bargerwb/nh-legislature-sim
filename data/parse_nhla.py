#!/usr/bin/env python3
"""
NHLA 2024 Rating Data → NH_DATA JSON for Legislature Simulator
Downloads nothing — expects 2024_rating_spreadsheet.xlsx in same directory.
"""

import json
import os
import re
import numpy as np
from collections import defaultdict

import openpyxl
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA

XLSX_PATH = os.path.join(os.path.dirname(__file__), '2024_rating_spreadsheet.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'nh_data.json')

NUM_ARCHETYPES = 6
NUM_BILL_TYPES = 5

ARCHETYPE_DEFS = [
    {"id": 0, "name": "Liberty Hawks",      "abbrev": "LH", "color": "#ffd60a"},
    {"id": 1, "name": "Fiscal Conservatives","abbrev": "FC", "color": "#3a86ff"},
    {"id": 2, "name": "Social Conservatives","abbrev": "SC", "color": "#fb5607"},
    {"id": 3, "name": "Establishment",       "abbrev": "ES", "color": "#9b5de5"},
    {"id": 4, "name": "Blue Dogs",           "abbrev": "BD", "color": "#06d6a0"},
    {"id": 5, "name": "Progressives",        "abbrev": "PR", "color": "#ff006e"},
]


def load_workbook():
    return openpyxl.load_workbook(XLSX_PATH, data_only=True)


def parse_house_info(wb):
    """Parse legislator info from House sheet."""
    ws = wb['House']
    legislators = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        full_name = str(row[0])
        party = str(row[4]) if row[4] else '?'
        county = str(row[5]) if row[5] else ''
        district = str(row[7]) if row[7] else ''
        score = float(row[18]) if row[18] is not None else 0
        grade = str(row[19]) if row[19] else '?'
        legislators.append({
            'fullName': full_name,
            'firstName': str(row[2]) if row[2] else '',
            'lastName': str(row[3]) if row[3] else '',
            'party': party,
            'county': county,
            'district': district,
            'score': round(score, 1),
            'grade': grade,
        })
    return legislators


def parse_vote_matrix(wb):
    """Parse House Reps RC sheet into vote matrix. Returns (names, rc_headers, matrix).
    Matrix values: 1=Correct, 0=Incorrect, NaN=absent/NA/presiding
    """
    ws = wb['House Reps RC']
    headers = [cell.value for cell in ws[1]]
    rc_headers = headers[2:]  # Skip Name, District

    names = []
    matrix = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        name = str(row[0])
        names.append(name)
        votes = []
        for v in row[2:]:
            sv = str(v).strip() if v else ''
            if sv == 'Correct':
                votes.append(1.0)
            elif sv == 'Incorrect':
                votes.append(0.0)
            else:
                votes.append(float('nan'))
        matrix.append(votes)

    return names, rc_headers, np.array(matrix)


def parse_house_votes(wb):
    """Parse bill info from HouseVotes sheet."""
    ws = wb['HouseVotes']
    bills = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        bill_number = str(row[0])
        title = str(row[1]) if row[1] else ''
        weight = float(row[2]) if row[2] is not None else 1
        motion = str(row[3]) if row[3] else ''
        correct_vote = str(row[4]) if row[4] else ''
        result = str(row[5]) if row[5] else ''
        rc_num = int(row[6]) if row[6] is not None else 0

        # Parse result into yeas/nays
        yeas, nays = 0, 0
        m = re.match(r'(\d+)-(\d+)', str(result))
        if m:
            yeas, nays = int(m.group(1)), int(m.group(2))

        # Determine NHLA position from correct vote
        nhla_position = 'support' if correct_vote == 'Y' else 'oppose'

        # Determine outcome
        passed = yeas > nays
        outcome = 'passed' if passed else 'failed'

        bills.append({
            'billNumber': bill_number,
            'title': title[:120],
            'weight': weight,
            'motion': motion,
            'correctVote': correct_vote,
            'nhlaPosition': nhla_position,
            'result': result,
            'yeas': yeas,
            'nays': nays,
            'outcome': outcome,
            'rcNum': rc_num,
        })
    return bills


def parse_sponsor_sentiment(wb):
    """Parse HouseSponsorData for bill sentiment."""
    ws = wb['HouseSponsorData']
    sentiments = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sentiment = str(row[1]) if row[1] else ''
        sentiments.append(sentiment)
    return sentiments


def impute_matrix(matrix):
    """Replace NaN with 0.5 (neutral) for clustering."""
    result = matrix.copy()
    result[np.isnan(result)] = 0.5
    return result


def assign_archetypes(matrix_imputed, n_clusters=NUM_ARCHETYPES):
    """K-means clustering on vote matrix to assign archetypes."""
    kmeans = KMeans(n_clusters=n_clusters, n_init=20, random_state=42, max_iter=300)
    labels = kmeans.fit_predict(matrix_imputed)
    centroids = kmeans.cluster_centers_

    # Sort clusters by mean NHLA alignment (descending) so cluster 0 = most libertarian
    cluster_means = [np.mean(centroids[i]) for i in range(n_clusters)]
    sorted_indices = np.argsort(cluster_means)[::-1]

    # Create remapping
    remap = {old: new for new, old in enumerate(sorted_indices)}
    remapped_labels = np.array([remap[l] for l in labels])
    remapped_centroids = centroids[sorted_indices]

    return remapped_labels, remapped_centroids


def compute_pca_positions(matrix_imputed):
    """PCA to 2D for particle positioning."""
    pca = PCA(n_components=2, random_state=42)
    coords = pca.fit_transform(matrix_imputed)

    # Normalize to [0.05, 0.95] range (fraction of canvas)
    for dim in range(2):
        mn, mx = coords[:, dim].min(), coords[:, dim].max()
        if mx > mn:
            coords[:, dim] = 0.05 + 0.9 * (coords[:, dim] - mn) / (mx - mn)
        else:
            coords[:, dim] = 0.5

    return coords


def compute_correlation_matrix(matrix_imputed, labels, n_clusters=NUM_ARCHETYPES):
    """Compute voting agreement between archetype pairs, scaled to [-1, +1]."""
    corr = np.zeros((n_clusters, n_clusters))

    for i in range(n_clusters):
        members_i = matrix_imputed[labels == i]
        mean_i = np.mean(members_i, axis=0)
        for j in range(n_clusters):
            members_j = matrix_imputed[labels == j]
            mean_j = np.mean(members_j, axis=0)
            # Agreement = correlation of mean vote profiles
            agreement = np.corrcoef(mean_i, mean_j)[0, 1]
            # Scale: same cluster ~1, opposite ~-1
            corr[i][j] = round(float(agreement), 3)

    return corr.tolist()


def compute_bill_archetype_reactions(matrix, labels, bills, rc_headers, n_clusters=NUM_ARCHETYPES):
    """For each bill (roll call), compute how each archetype voted."""
    # Map RC# to column index
    rc_to_col = {}
    for idx, header in enumerate(rc_headers):
        m = re.match(r'H-(\d+)', str(header))
        if m:
            rc_to_col[int(m.group(1))] = idx

    for bill in bills:
        rc_num = bill['rcNum']
        col_idx = rc_to_col.get(rc_num)
        if col_idx is None:
            bill['archetypeReaction'] = [0.0] * n_clusters
            continue

        col_data = matrix[:, col_idx]
        reactions = []
        for arch in range(n_clusters):
            mask = labels == arch
            arch_votes = col_data[mask]
            valid = arch_votes[~np.isnan(arch_votes)]
            if len(valid) == 0:
                reactions.append(0.0)
            else:
                # Mean vote: 1=correct(aligned), 0=incorrect
                # Scale to [-1, +1]: mean 1.0 → +1, mean 0.0 → -1, mean 0.5 → 0
                mean_vote = np.mean(valid)
                reaction = round(float(mean_vote * 2 - 1), 3)
                reactions.append(reaction)
        bill['archetypeReaction'] = reactions

    return bills


def cluster_bills(bills, n_bill_types=NUM_BILL_TYPES):
    """Cluster bills by their archetype reaction vectors into bill personality types."""
    reaction_vectors = np.array([b['archetypeReaction'] for b in bills])

    # Some bills might have all-zero reactions — still include them
    kmeans = KMeans(n_clusters=n_bill_types, n_init=20, random_state=42, max_iter=300)
    bill_labels = kmeans.fit_predict(reaction_vectors)
    centroids = kmeans.cluster_centers_

    # Sort by mean reaction (most pro-liberty first)
    cluster_means = [np.mean(centroids[i]) for i in range(n_bill_types)]
    sorted_indices = np.argsort(cluster_means)[::-1]
    remap = {old: new for new, old in enumerate(sorted_indices)}
    remapped_labels = np.array([remap[l] for l in bill_labels])
    remapped_centroids = centroids[sorted_indices]

    # Name bill types based on their reaction profiles
    bill_type_names = []
    used_names = set()
    for i in range(n_bill_types):
        centroid = remapped_centroids[i]
        mean_r = np.mean(centroid)

        # Check for cross-party signature: LH agrees with BD/PR while FC/SC disagree
        lh_r = centroid[0] if len(centroid) > 0 else 0
        fc_r = centroid[1] if len(centroid) > 1 else 0
        sc_r = centroid[2] if len(centroid) > 2 else 0
        bd_r = centroid[4] if len(centroid) > 4 else 0
        pr_r = centroid[5] if len(centroid) > 5 else 0
        r_block = (lh_r + fc_r + sc_r) / 3
        d_block = (bd_r + pr_r) / 2

        if min(centroid) > 0.5:
            name = "Bipartisan"
        elif lh_r > 0.2 and pr_r > 0.2 and fc_r < 0:
            name = "Personal Freedom"
        elif r_block > 0.5 and d_block < -0.5:
            if mean_r > 0.2:
                name = "Partisan Liberty"
            else:
                name = "Hard Partisan"
        elif mean_r > 0.1:
            name = "Moderate Liberty"
        elif r_block < 0 and d_block > 0:
            name = "Statist"
        else:
            name = "Divisive Split"

        # Make unique
        if name in used_names:
            count = sum(1 for n in bill_type_names if n.startswith(name))
            # Don't number the first occurrence, number the second
            old_idx = bill_type_names.index(name)
            bill_type_names[old_idx] = f"{name} I"
            name = f"{name} II"
        used_names.add(name)
        bill_type_names.append(name)

    # Assign bill types
    for idx, bill in enumerate(bills):
        bill['billType'] = int(remapped_labels[idx])

    bill_type_colors = ['#22c55e', '#84cc16', '#eab308', '#f97316', '#ef4444']

    bill_types = []
    for i in range(n_bill_types):
        bill_types.append({
            'id': i,
            'name': bill_type_names[i],
            'centroidReaction': [round(float(v), 3) for v in remapped_centroids[i]],
            'color': bill_type_colors[i % len(bill_type_colors)],
        })

    return bill_types, bills


def compute_bill_positions(bills, pca_coords, labels):
    """Compute bill position as weighted average of supporter positions."""
    for bill in bills:
        reactions = bill['archetypeReaction']
        # Weight each legislator's position by their archetype's reaction
        wx_sum, wy_sum, w_sum = 0, 0, 0
        for idx in range(len(pca_coords)):
            arch = labels[idx]
            reaction = reactions[arch]
            weight = max(0, reaction)  # Only attracted legislators
            wx_sum += pca_coords[idx][0] * weight
            wy_sum += pca_coords[idx][1] * weight
            w_sum += weight

        if w_sum > 0:
            bill['bx'] = round(float(wx_sum / w_sum), 4)
            bill['by'] = round(float(wy_sum / w_sum), 4)
        else:
            bill['bx'] = 0.5
            bill['by'] = 0.5

    return bills


def main():
    print("Loading workbook...")
    wb = load_workbook()

    print("Parsing legislator info...")
    legislators = parse_house_info(wb)
    print(f"  Found {len(legislators)} legislators")

    print("Parsing vote matrix...")
    rc_names, rc_headers, vote_matrix = parse_vote_matrix(wb)
    print(f"  Matrix: {vote_matrix.shape[0]} legislators x {vote_matrix.shape[1]} roll calls")

    print("Parsing bill info...")
    bills = parse_house_votes(wb)
    print(f"  Found {len(bills)} bill votes")

    print("Imputing missing votes...")
    matrix_imputed = impute_matrix(vote_matrix)

    print(f"Running K-means (k={NUM_ARCHETYPES}) for legislator archetypes...")
    labels, centroids = assign_archetypes(matrix_imputed, NUM_ARCHETYPES)
    for i in range(NUM_ARCHETYPES):
        count = np.sum(labels == i)
        mean_score = np.mean(matrix_imputed[labels == i])
        print(f"  Archetype {i} ({ARCHETYPE_DEFS[i]['name']}): {count} members, mean alignment: {mean_score:.3f}")

    print("Computing PCA positions...")
    pca_coords = compute_pca_positions(matrix_imputed)

    print("Computing correlation matrix...")
    corr_matrix = compute_correlation_matrix(matrix_imputed, labels, NUM_ARCHETYPES)

    print("Computing bill archetype reactions...")
    bills = compute_bill_archetype_reactions(vote_matrix, labels, bills, rc_headers, NUM_ARCHETYPES)

    print(f"Clustering bills (k={NUM_BILL_TYPES})...")
    bill_types, bills = cluster_bills(bills, NUM_BILL_TYPES)
    for bt in bill_types:
        count = sum(1 for b in bills if b['billType'] == bt['id'])
        print(f"  Bill type {bt['id']} ({bt['name']}): {count} bills, centroid: {bt['centroidReaction']}")

    print("Computing bill positions...")
    bills = compute_bill_positions(bills, pca_coords, labels)

    # Match legislators to their vote matrix row by name
    # Build name lookup: rc_names are like "Abare,Kimberly", legislators have lastName/firstName
    rc_name_to_idx = {}
    for idx, name in enumerate(rc_names):
        rc_name_to_idx[name.strip().lower()] = idx

    # Assign archetype + PCA coords to legislators
    matched = 0
    for leg in legislators:
        # Try matching by "LastName,FirstName"
        lookup = f"{leg['lastName']},{leg['firstName']}".lower()
        idx = rc_name_to_idx.get(lookup)
        if idx is None:
            # Try without spaces
            lookup2 = lookup.replace(' ', '')
            for key, val in rc_name_to_idx.items():
                if key.replace(' ', '') == lookup2:
                    idx = val
                    break
        if idx is not None:
            leg['archetype'] = int(labels[idx])
            leg['px'] = round(float(pca_coords[idx][0]), 4)
            leg['py'] = round(float(pca_coords[idx][1]), 4)
            matched += 1
        else:
            # Default: assign based on score
            if leg['score'] >= 87:
                leg['archetype'] = 0
            elif leg['score'] >= 70:
                leg['archetype'] = 1
            elif leg['score'] >= 50:
                leg['archetype'] = 2
            elif leg['score'] >= 30:
                leg['archetype'] = 3
            elif leg['score'] >= 15:
                leg['archetype'] = 4
            else:
                leg['archetype'] = 5
            leg['px'] = 0.5 + (np.random.random() - 0.5) * 0.2
            leg['py'] = 0.5 + (np.random.random() - 0.5) * 0.2

    print(f"  Matched {matched}/{len(legislators)} legislators to vote data")

    # Clean up bill objects for JSON
    bills_clean = []
    for i, b in enumerate(bills):
        bills_clean.append({
            'id': i,
            'billNumber': b['billNumber'],
            'title': b['title'],
            'motion': b['motion'],
            'nhlaPosition': b['nhlaPosition'],
            'outcome': b['outcome'],
            'billType': b['billType'],
            'archetypeReaction': b['archetypeReaction'],
            'bx': b['bx'],
            'by': b['by'],
            'yeas': b['yeas'],
            'nays': b['nays'],
            'weight': b['weight'],
        })

    # Build NH_DATA
    nh_data = {
        'meta': {
            'year': 2024,
            'chamber': 'house',
            'totalBills': len(bills_clean),
            'totalLegislators': len(legislators),
        },
        'archetypes': ARCHETYPE_DEFS,
        'correlationMatrix': corr_matrix,
        'billTypes': bill_types,
        'legislators': legislators,
        'bills': bills_clean,
    }

    print(f"\nWriting {OUTPUT_PATH}...")
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(nh_data, f, separators=(',', ':'))

    # Also write a pretty version for debugging
    debug_path = OUTPUT_PATH.replace('.json', '_debug.json')
    with open(debug_path, 'w') as f:
        json.dump(nh_data, f, indent=2)

    file_size = os.path.getsize(OUTPUT_PATH)
    print(f"Done! Output: {file_size:,} bytes")
    print(f"  {len(legislators)} legislators, {len(bills_clean)} bills")
    print(f"  {NUM_ARCHETYPES} archetypes, {NUM_BILL_TYPES} bill types")
    print(f"  Correlation matrix: {len(corr_matrix)}x{len(corr_matrix[0])}")


if __name__ == '__main__':
    main()
